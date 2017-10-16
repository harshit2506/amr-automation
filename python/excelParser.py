from openpyxl import load_workbook
import datetime

constant_dict = {
    'assignment_team_value': 'SUPPLIERNET',
    'closed_status': 'CLOSED',
    'file_name' : 'P11W01 AMR ALL',
    'path_to_source' : '..\\dummy source excel.xlsx',
    'data_sheet_name': 'Sheet1',
    'path_to_destination': '..\\',
    'destination_file_name':'dummy destination file.xlsx',
    'destination_worksheet': 'Sheet1',
    'output_file_name': 'output.xlsx,
    'portfolio_value' : 'FLNA',
    'cluster_group' : 'Finance',
    'cluster' : 'Track1',
    'effort_bucket' : 'LE40',
    'category_in_destination' : 'User Support',
    'sub_category_in_destination' : 'User Maintenance'
}
column_idx_dict = {
    'assignment_team': 6,
    'status': 5,
    'open_date' : 16,
    'close_period': 18,
    'close_week': 19,
    'close_year': 17,
    'open_year': 9,
    'open_period': 10,
    'open_week': 11,
    'urgency': 3,
    'environment': 4,
    'request_number': 1,
    'committed_date': 31,
    'description': 40,
    'category' : 38
}

urgency_mapping_dict={
    '1' : 'Emergency',
    '2' : 'High',
    '3' : 'Medium',
    '4' : 'Low'
}

source_target_mapping_dict={
    'application' : 'category'
}

source_workbook = load_workbook(filename=constant_dict['path_to_source'], read_only=True)
source_worksheet = source_workbook[constant_dict['data_sheet_name']]

def check_column_indexes():
    header_row_idx = 1
    for each_key in column_idx_dict.keys():
        print(each_key, ':' ,source_worksheet[header_row_idx][column_idx_dict[each_key]].value)

def check_closed_conditions(current_row):
    return_val = False
    if current_row[column_idx_dict['close_week']].value.strip() != '':
        if int(current_row[column_idx_dict['close_week']].value.strip()) == int(constant_dict['file_name'][4:6]):
            if int(current_row[column_idx_dict['close_period']].value.strip()) == int(constant_dict['file_name'][1:3]):
                if int(current_row[column_idx_dict['close_year']].value.strip()) == datetime.datetime.now().year:
                    return_val = True
    return return_val

def check_open_conditions(current_row):
    return_val = False
    if current_row[column_idx_dict['close_week']].value.strip() == '':
        if int(current_row[column_idx_dict['open_week']].value.strip()) == int(constant_dict['file_name'][4:6]):
            if int(current_row[column_idx_dict['open_period']].value.strip()) == int(constant_dict['file_name'][1:3]):
                if int(current_row[column_idx_dict['open_year']].value.strip()) == datetime.datetime.now().year:
                    return_val = True
    return return_val


def extract_data():
    result_data_list = []
    for row_idx in range(1, (source_worksheet.max_row + 1)):
        if source_worksheet[row_idx][column_idx_dict['assignment_team']].value == constant_dict['assignment_team_value']:
            if check_closed_conditions(source_worksheet[row_idx]) or check_open_conditions(source_worksheet[row_idx]):
                result_data_list.append(source_worksheet[row_idx])
    return result_data_list

def update_target_excel(result_list):
    target_workbook = load_workbook(constant_dict['path_to_destination'] + constant_dict['destination_file_name'])
    target_worksheet = target_workbook[constant_dict['destination_worksheet']]
    # TODO: put logic
    start_idx = target_worksheet.max_row + 1
    for each_row in result_list:
        if each_row[column_idx_dict['status']].value.upper() == constant_dict['closed_status'] and update_if_present(each_row):
            continue
        result_row_list = [constant_dict['portfolio_value'], constant_dict['cluster_group'], constant_dict['cluster']]
        result_row_list.append(each_row[column_idx_dict['request_number']].value)
        result_row_list.append(urgency_mapping_dict[each_row[column_idx_dict['urgency']].value])
        result_row_list.append(each_row[column_idx_dict['status']].value)
        result_row_list.append(each_row[column_idx_dict['environment']].value)
        result_row_list.append(each_row[column_idx_dict[source_target_mapping_dict['application']]].value)
        result_row_list.append(each_row[column_idx_dict['description']].value)
        result_row_list.append(constant_dict['category_in_destination'])
        result_row_list.append(constant_dict['sub_category_in_destination'])
        result_row_list.append(constant_dict['effort_bucket'])
        result_row_list.append(each_row[column_idx_dict['open_date']].value)
        result_row_list.append(each_row[column_idx_dict['open_year']].value)
        result_row_list.append('P' + int(each_row[column_idx_dict['open_period']].value))
        result_row_list.append('W' + int(each_row[column_idx_dict['open_week']].value))
        result_row_list.append(each_row[column_idx_dict['open_period']].value))
        result_row_list.append("")
        result_row_list.append(each_row[column_idx_dict['committed_date']].value)
        result_row_list.append("")
        result_row_list.append("")
        result_row_list.append(each_row[column_idx_dict['open_date']].value)
        result_row_list.append(each_row[column_idx_dict['open_year']].value)
        result_row_list.append('P' + int(each_row[column_idx_dict['open_period']].value))
        result_row_list.append('W' + int(each_row[column_idx_dict['open_week']].value))
        result_row_list.append(each_row[column_idx_dict['open_date']].value)
        result_row_list.append(each_row[column_idx_dict['open_year']].value)
        result_row_list.append('P' + int(each_row[column_idx_dict['open_period']].value))
        result_row_list.append('W' + int(each_row[column_idx_dict['open_week']].value))




    target_workbook.save(constant_dict['path_to_destination'] + constant_dict['output_file_name'])

update_target_excel(None)
